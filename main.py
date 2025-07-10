# код был написан https://t.me/Mjoer1

import os
import requests
import json
from openpyxl import Workbook
from datetime import datetime, timedelta
from time import sleep


# введите токен личного кабинета ВБ (получить его можно по ссылке https://seller.wildberries.ru/supplier-settings/access-to-api)
token = ''


if token == '':
    print('Введите токен личного кабинета ВБ')

headers = {'Authorization': token, "Content-type": 'application/json'}


# функция для проверки авторизации по токену
def test():
    r = requests.get(url="https://common-api.wildberries.ru/ping", headers=headers)
    if r.status_code == 200:
        return 'Авторизация произошла успешно'
    elif r.status_code == 401:
        return 'Какие-то проблемы с токеном. Перепроверьте корректность введеного токена'
    elif r.status_code == 429:
        return 'Подождите пару минут и повторите запрос'
    else:
        return 'Неизвестная ошибка'


# функция для инициализации товаров и данных о них (себестоимость, наименование, основной склад, стоимость логистики)
def init_goods():
    goods_dict = {}
    if os.path.exists('goods.json'):
        with open('goods.json', 'r') as fr:
            goods_dict = json.load(fr)
    else:

        flag = 1
        while flag:
            articul = input('Введите артикул товара (тот же, что и в отчете ВБ). Если добавление товаров окончено, то введите 0\n')
            if articul == '0':
                flag = 0
                break
            goods_dict[articul] = {}
            name = input('Введите наименование товара\n')
            price = float(input('Введите себестоимость товара без учета логистики на склад ВБ транспортной комппаней\n'))
            main_warehouse = input('Введите название основного склада ВБ для данного товара (доставка на который будет стоить 0 рублей)\n')
            logistics_price = float(input('Введите стоимость транспортировки товара на любой другой склад (либо же среднее значение этой стоимости)\n'))
            print()

            goods_dict[articul]['наименование'] = name
            goods_dict[articul]['себестоимость'] = price
            goods_dict[articul]['основной склад'] = main_warehouse
            goods_dict[articul]['стоимость транспортировки'] = logistics_price
        with open("goods.json", "w") as fw:
            json.dump(goods_dict, fw, ensure_ascii=False)
        print('Отредактировать значения вручную можно в файле goods.json\n')
    return goods_dict


# функция для получения недельных отчетов о продажах по реализации за выбранный период
# принимает начальную дату отчета, конечную дату отчета в формате datetime
def get_week_report(start_date: datetime, end_date: datetime = datetime.today()):
    params = {
        'dateFrom': start_date.strftime('%Y-%m-%d'),
        'dateTo': end_date.strftime('%Y-%m-%d')
    }
    r = requests.get(url="https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod", headers=headers, params=params)
    if r.status_code == 200:
        return r.json()
    if r.status_code == 401:
        return test()
    if r.status_code == 429:
        return 'Слишком много запросов. Подождите несколько минут и повторите попытку'


# функция для анализа еженедельных отчетов и создания файла эксель из этих отчетов
# принимает отчет в формате списка, имя файла, в которой запишем результат и имя листа таблицы в эксель
def create_excel_week_report(dict_list: list, output_filename: str, sheet_name='Недельный отчет'):
    # Создаем директорию, если она не существует
    if not os.path.exists('week_report'):
        os.makedirs('week_report')

    filepath = os.path.join('week_report', output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # словарь для хранения данных о каждой позиции: ключ - артикул, значение - словарь с данными о каждой позиции
    res_tovar = {}

    # Записываем данные из списка словарей в Excel
    if dict_list:
        # Записываем заголовки
        ws.append(['Дата продажи', 'Категория', 'Артикул', 'Кол-во проданных', 'Кол-во возвратов',
                   'Цена реализации (указываем как доход)', 'Перечислено продавцу', 'Стоимость логистики', 'Склад доставки', 'Коэффииент склада', 'Траты на рекламу'])

        for realization in dict_list:
            sale_date = realization['rr_dt']
            category = realization['subject_name']
            articul = realization['sa_name']
            amount_sale = realization['quantity']
            amount_return = realization['return_amount']
            realize_price = realization['retail_amount']
            clean_price = realization['ppvz_for_pay']
            delivery_price = realization['delivery_rub']
            office_name = realization['office_name']
            office_coef = realization['dlv_prc']
            add_price = realization['deduction']
            if add_price:
                if 'deduction' in res_tovar:
                    res_tovar['deduction'] += add_price
                else:
                    res_tovar['deduction'] = add_price


            ws.append([sale_date, category, articul, amount_sale, amount_return, realize_price,
                       clean_price, delivery_price, office_name, office_coef, add_price])
            if articul:
                if articul in res_tovar:
                    res_tovar[articul]['amount_sale'] += amount_sale
                    res_tovar[articul]['amount_return'] += amount_return
                    res_tovar[articul]['realize_price'] += realize_price
                    res_tovar[articul]['clean_price'] += clean_price
                    res_tovar[articul]['delivery_price'] += delivery_price
                    if office_name != goods_info_dict[articul]['основной склад']:
                        res_tovar[articul]['delivery_price'] += goods_info_dict[articul]['стоимость транспортировки']
                else:
                    res_tovar[articul] = {'amount_sale': amount_sale, 'amount_return': amount_return, 'realize_price': realize_price,
                                            'clean_price': clean_price, 'delivery_price': delivery_price}
    wb.save(filepath)
    return res_tovar


# функция для получения отчета о тратах на платное хранение за выбранный срок
# принимает начальную дату отчета, конечную дату отчета в формате datetime
def get_paid_storage_report(start_date: datetime = datetime.today(), end_date: datetime = datetime.today()):
    flag = 1
    mistake_flag = 0
    res_pay = {}
    while flag:
        if end_date - timedelta(days=6) > start_date:
            end_end_date = end_date
            end_date = start_date + timedelta(days=6)
        else:
            break
        params = {
            'dateFrom': start_date.strftime('%Y-%m-%d'),
            'dateTo': end_date.strftime('%Y-%m-%d')
        }

        report_id = requests.get(url='https://seller-analytics-api.wildberries.ru/api/v1/paid_storage', headers=headers,
                                 params=params)
        if report_id.status_code == 200:
            report_id = report_id.json()['data']['taskId']
        else:
            print(f'Ошибка получения отчета о платном хранении. ID ошибки {report_id.status_code}')

        for i in range(5):
            sleep(3)
            r_check_report_status = requests.get(
                url=f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{report_id}/status',
                headers=headers).json()
            if r_check_report_status['data']['status'] == 'done':
                break

        report = requests.get(
            url=f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{report_id}/download',
            headers=headers).json()

        for pay in report:
            try:
                name = pay['vendorCode']  # артикул товара
                amount = pay['warehousePrice']  # плата за хранение товара
            except TypeError:
                print('Ошибка прогрузки данных о платном хранении')
                sleep(65)
                mistake_flag = 1
                break
            res_pay[name] = res_pay.get(name, 0) + amount
        if mistake_flag:
            mistake_flag = 0
            report = requests.get(
                url=f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{report_id}/download',
                headers=headers).json()

            for pay in report:
                try:
                    name = pay['vendorCode']  # артикул товара
                    amount = pay['warehousePrice']  # плата за хранение товара
                except TypeError:
                    print('Ошибка прогрузки данных о платном хранении')
                    sleep(65)
                    break
                res_pay[name] = res_pay.get(name, 0) + amount
        start_date = end_date + timedelta(days=1)
        end_date = end_end_date

    sleep(65)
    if end_date > start_date:
        params = {
            'dateFrom': start_date.strftime('%Y-%m-%d'),
            'dateTo': end_date.strftime('%Y-%m-%d')
        }

        report_id = requests.get(url='https://seller-analytics-api.wildberries.ru/api/v1/paid_storage', headers=headers,
                                 params=params)
        if report_id.status_code == 200:
            report_id = report_id.json()['data']['taskId']
        else:
            print(f'Ошибка получения отчета о платном хранении. ID ошибки {report_id.status_code}')

        for i in range(5):
            sleep(3)
            r_check_report_status = requests.get(
                url=f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{report_id}/status',
                headers=headers).json()
            if r_check_report_status['data']['status'] == 'done':
                break

        report = requests.get(
            url=f'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{report_id}/download',
            headers=headers).json()

        for pay in report:
            try:
                name = pay['vendorCode']  # артикул товара
                amount = pay['warehousePrice']  # плата за хранение товара
            except TypeError:
                print('Ошибка прогрузки данных о платном хранении, повторите запрос через несколько минут')
                return res_pay

            res_pay[name] = res_pay.get(name, 0) + amount
    return res_pay


# функция, которая создает таблицу в эксель с прибылью
def make_excel_final_report(res_report, res_storage, start_date, end_date):
    # Создаем директорию, если она не существует
    if not os.path.exists('week_report'):
        os.makedirs('week_report')

    filepath = os.path.join('week_report', f'Итоговый отчет {start_date.strftime("%d.%m.%Y")}-{end_date.strftime("%d.%m.%Y")}.xlsx')

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчет'

    # Записываем данные из списка словарей в Excel
    # Записываем заголовки
    ws.append(['Артикул', 'Наименование', 'Кол-во проданных', 'Кол-во возвратов', 'Цена реализации (указываем как доход)',
               'Перечислено продавцу', 'Стоимость логистики', 'Траты на хранение', 'Траты на рекламу',
               'Налог 1%', 'Себестоимость товара', 'Прибыль (до налога 15%)', 'Итоговая прибыль', 'Прибыль за штуку', 'Логистика за штуку'])

    try:
        ws.append(['', '', '', '', '', '', '', '', res_report['deduction']])
    except:
        ws.append(['', '', '', '', '', '', '', '', 0])
    ws.append([])

    for k, v in res_report.items():
        if k != 'deduction':
            try:
                res_storage[k]
            except KeyError:
                res_storage[k] = 0
            ws.append([k, goods_info_dict[k]['наименование'], v['amount_sale'], v['amount_return'], v['realize_price'], v['clean_price'],
                   v['delivery_price'], res_storage[k], 0, v['realize_price']*0.01, v['amount_sale']*goods_info_dict[k]['себестоимость'], 0])
            ws.append([])

    for row in range(4, ws.max_row+1, 2):
        ws.cell(row=row, column=12).value = f"=F{row} - G{row} - H{row} - I{row} - J{row} - K{row}"
        articul = ws.cell(row=row, column=1).value
        profit =  (res_report[articul]['clean_price'] - res_report[articul]['realize_price']*0.01 -
                   res_report[articul]['delivery_price'] - res_report[articul]['amount_sale']*goods_info_dict[articul]['себестоимость']
                   - res_storage[articul])
        if profit > 0:
            ws.cell(row=row, column=13).value = f"=L{row}*0.85"
        else:
            ws.cell(row=row, column=13).value = f"=L{row}"
        # прибыль за штуку
        ws.cell(row=row, column=14).value = f"=M{row}/C{row}"
        # логистика за штуку
        ws.cell(row=row, column=15).value = f"=G{row}/C{row}"
    wb.save(filepath)
    print('Итоговый отчет составлен')

# функция которая создает финальный отчет о прибыли, соединяя в себе ранее написанные функции
# создает таблицу в эксель
def make_final_report(start_date=datetime(year=2025, month=1, day=13), end_date=datetime(year=2025, month=1, day=19)):
    res_storage = get_paid_storage_report(start_date=start_date, end_date=end_date)


    week_report_json = get_week_report(start_date=start_date, end_date=end_date)

    res_report = create_excel_week_report(dict_list=week_report_json,
                                          output_filename=f'Отчет {start_date.strftime("%d.%m.%Y")}-{end_date.strftime("%d.%m.%Y")}.xlsx')

    make_excel_final_report(res_report, res_storage, start_date, end_date)



if __name__ == "__main__":
    print(test())
    goods_dict = init_goods()

    goods_info_dict = init_goods()

    make_final_report(start_date=datetime(year=2025, month=6, day=30), end_date=datetime(year=2025, month=7, day=6))